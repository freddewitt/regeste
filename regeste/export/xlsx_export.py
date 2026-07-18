"""XLSX export — one tab per fonds/série plus an overview tab, conditional formatting on OCR
confidence, auto-filters, and a thumbnail image per row.
"""

from __future__ import annotations

import re
from pathlib import Path

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from regeste.pivot import Piece, global_status

from .common import filter_pieces, hierarchy_path
from .xlsx_common import (
    apply_confidence_conditional_formatting,
    autofilter,
    insert_thumbnail,
    write_header,
)

HEADERS = [
    "id",
    "call_number",
    "date",
    "sender",
    "recipient",
    "summary",
    "confidence_score",
    "status",
    "image",
]


def export_xlsx(
    pieces: list[Piece],
    output_path: Path,
    *,
    validated_only: bool = False,
    target_language: str | None = None,
) -> Path:
    pieces = filter_pieces(pieces, validated_only=validated_only)
    wb = Workbook()
    overview = wb.active
    overview.title = "Vue d'ensemble"
    _write_sheet(overview, pieces, target_language)

    groups: dict[str, list[Piece]] = {}
    for piece in pieces:
        key = " - ".join(hierarchy_path(piece)) or "Sans hiérarchie"
        groups.setdefault(key, []).append(piece)
    for key, group_pieces in groups.items():
        title = _sanitize_sheet_title(key)
        sheet = wb.create_sheet(title=title)
        _write_sheet(sheet, group_pieces, target_language)

    wb.save(output_path)
    return output_path


_INVALID_SHEET_TITLE_CHARS = re.compile(r"[\\*?:/\[\]]")


def _sanitize_sheet_title(title: str) -> str:
    cleaned = _INVALID_SHEET_TITLE_CHARS.sub("-", title)[:31]
    return cleaned or "Sans hiérarchie"


def _write_sheet(sheet: Worksheet, pieces: list[Piece], target_language: str | None) -> None:
    headers = [*HEADERS, "translation"] if target_language else HEADERS
    write_header(sheet, headers)
    for row, piece in enumerate(pieces, start=2):
        sheet.cell(row=row, column=1, value=piece.id)
        sheet.cell(row=row, column=2, value=piece.call_number)
        sheet.cell(row=row, column=3, value=piece.date)
        sheet.cell(row=row, column=4, value=piece.sender)
        sheet.cell(row=row, column=5, value=piece.recipient)
        sheet.cell(row=row, column=6, value=piece.summary)
        sheet.cell(row=row, column=7, value=piece.confidence_score)
        sheet.cell(row=row, column=8, value=global_status(piece))
        if piece.image_path:
            insert_thumbnail(sheet, Path(piece.image_path), f"I{row}")
        if target_language:
            translation = (piece.translations or {}).get(target_language)
            sheet.cell(row=row, column=10, value=translation.text if translation else "")
    if pieces:
        apply_confidence_conditional_formatting(sheet, "G", 2, len(pieces) + 1)
        autofilter(sheet)
