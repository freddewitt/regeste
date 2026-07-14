"""Corpus-level "journal de revue" XLSX export — one row per piece, aggregated across every
pivot JSON in the corpus. Distinct from the per-piece exporters (`export/xlsx_export.py`)
but reuses the same XLSX helpers (`export/xlsx_common.py`).
"""

from __future__ import annotations

from datetime import datetime, timezone
from pathlib import Path

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from regeste.pivot import CONTENT_FIELDS, FieldValidation, Piece, global_status

from .common import hierarchy_path
from .xlsx_common import apply_confidence_conditional_formatting, autofilter, write_header

DETAIL_HEADERS = [
    "id",
    "call_number",
    "fonds",
    "series",
    *[f"status_{field}" for field in CONTENT_FIELDS],
    "global_status",
    "confidence_score",
    "validated_by",
    "validated_at",
    "translation_languages",
]

SUMMARY_HEADERS = ["fonds/série", "total", "% validé", "% à vérifier/brouillon", "% rejeté"]


def export_review_journal(
    pieces: list[Piece], output_path: Path, *, generated_at: str | None = None
) -> Path:
    generated_at = generated_at or datetime.now(timezone.utc).isoformat()

    wb = Workbook()
    detail = wb.active
    detail.title = "Détail"
    detail.cell(row=1, column=1, value=f"Rapport généré le {generated_at}")
    write_header(detail, DETAIL_HEADERS, row=2)
    for row, piece in enumerate(pieces, start=3):
        _write_detail_row(detail, row, piece)
    if pieces:
        confidence_col = get_column_letter(DETAIL_HEADERS.index("confidence_score") + 1)
        apply_confidence_conditional_formatting(detail, confidence_col, 3, len(pieces) + 2)
        autofilter(detail, header_row=2)

    summary = wb.create_sheet(title="Synthèse")
    write_header(summary, SUMMARY_HEADERS)
    for row, (key, counts) in enumerate(_summarize_by_hierarchy(pieces).items(), start=2):
        total = sum(counts.values())
        summary.cell(row=row, column=1, value=key)
        summary.cell(row=row, column=2, value=total)
        summary.cell(row=row, column=3, value=_pct(counts.get("validated", 0), total))
        summary.cell(
            row=row,
            column=4,
            value=_pct(counts.get("to_review", 0) + counts.get("draft", 0), total),
        )
        summary.cell(row=row, column=5, value=_pct(counts.get("rejected", 0), total))

    wb.save(output_path)
    return output_path


def _write_detail_row(sheet: Worksheet, row: int, piece: Piece) -> None:
    col = 1
    for value in (piece.id, piece.call_number, piece.fonds, piece.series):
        sheet.cell(row=row, column=col, value=value)
        col += 1
    for field in CONTENT_FIELDS:
        status = piece.field_validations.get(field, FieldValidation())
        sheet.cell(row=row, column=col, value=status.status)
        col += 1
    sheet.cell(row=row, column=col, value=global_status(piece))
    col += 1
    sheet.cell(row=row, column=col, value=piece.confidence_score)
    col += 1
    latest = _latest_validation(piece)
    sheet.cell(row=row, column=col, value=latest.validated_by if latest else None)
    col += 1
    sheet.cell(row=row, column=col, value=latest.validated_at if latest else None)
    col += 1
    sheet.cell(row=row, column=col, value=",".join(sorted((piece.translations or {}).keys())))


def _latest_validation(piece: Piece) -> FieldValidation | None:
    validations = [v for v in piece.field_validations.values() if v.validated_at]
    return max(validations, key=lambda v: v.validated_at, default=None)


def _summarize_by_hierarchy(pieces: list[Piece]) -> dict[str, dict[str, int]]:
    summary: dict[str, dict[str, int]] = {}
    for piece in pieces:
        key = " - ".join(hierarchy_path(piece)) or "Sans hiérarchie"
        counts = summary.setdefault(key, {})
        status = global_status(piece)
        counts[status] = counts.get(status, 0) + 1
    return summary


def _pct(count: int, total: int) -> float:
    return round(100 * count / total, 1) if total else 0.0
