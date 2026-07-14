"""Shared XLSX helpers — conditional formatting, auto-filters, thumbnails.

Used by `export/xlsx_export.py` (per-piece) and `export/journal.py` (corpus-level).
"""

from __future__ import annotations

from pathlib import Path

from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet


def write_header(ws: Worksheet, headers: list[str], *, row: int = 1) -> None:
    for col, header in enumerate(headers, start=1):
        ws.cell(row=row, column=col, value=header)
    ws.freeze_panes = ws.cell(row=row + 1, column=1).coordinate


def autofilter(ws: Worksheet, *, header_row: int = 1) -> None:
    if ws.max_row <= header_row:
        return
    ws.auto_filter.ref = f"A{header_row}:{get_column_letter(ws.max_column)}{ws.max_row}"


def apply_confidence_conditional_formatting(
    ws: Worksheet, column_letter: str, first_row: int, last_row: int
) -> None:
    """Red (low confidence) to green (high confidence) color scale."""
    if last_row < first_row:
        return
    ws.conditional_formatting.add(
        f"{column_letter}{first_row}:{column_letter}{last_row}",
        ColorScaleRule(
            start_type="min",
            start_color="F8696B",
            mid_type="percentile",
            mid_value=50,
            mid_color="FFEB84",
            end_type="max",
            end_color="63BE7B",
        ),
    )


def insert_thumbnail(ws: Worksheet, image_path: Path, cell: str, *, max_size: int = 80) -> None:
    """Insert a small thumbnail at `cell` if `image_path` exists; skipped silently otherwise."""
    if not image_path.exists():
        return
    import io

    from openpyxl.drawing.image import Image as XLImage
    from PIL import Image as PILImage

    with PILImage.open(image_path) as source:
        thumb = source.convert("RGB")
        thumb.thumbnail((max_size, max_size))
        buffer = io.BytesIO()
        thumb.save(buffer, format="PNG")
    buffer.seek(0)
    ws.add_image(XLImage(buffer), cell)
