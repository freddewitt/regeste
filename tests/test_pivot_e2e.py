"""End-to-end verification of every export_instruct.md acceptance criterion, on a single
shared corpus that goes through the full lifecycle: OCR registry -> pivot build -> review
(validation, rejection, bulk-validate) -> translation -> export (12 formats + journal).
"""

from __future__ import annotations

import sqlite3
import xml.etree.ElementTree as ET
import zipfile

import pytest
from openpyxl import load_workbook
from PIL import Image as PILImage

from regeste.core.registry import FileEntry, Registry
from regeste.export import (
    FIELD_MAPPING,
    export_csv_light,
    export_dublin_core,
    export_ead,
    export_html,
    export_markdown,
    export_markdown_obsidian,
    export_mets,
    export_pdf,
    export_review_journal,
    export_sqlite,
    export_xlsx,
    export_zip,
)
from regeste.pivot import FieldValidation, build_pieces_from_registry, global_status, load_piece, save_piece
from regeste.review import apply_correction, apply_field_validation, bulk_validate, sorted_by_confidence
from regeste.translation import TranslationBlocked, TranslationProvider, TranslationResult, translate_piece


class _FakeTranslationProvider(TranslationProvider):
    name = "fake"

    def __init__(self) -> None:
        self.received_prompt = None

    @property
    def requires_api_key(self) -> bool:
        return False

    def translate(self, prompt: str, *, model: str) -> TranslationResult:
        self.received_prompt = prompt
        return TranslationResult(text="Dear Sir, I write from Lyon.", tokens_in=1, tokens_out=1, model=model)


@pytest.fixture
def corpus(tmp_path):
    """3 pieces seeded from a real OCR registry: fonds/série hierarchy, one destined to be
    fully validated + translated, one partially validated, one untouched (draft).
    """
    source_dir = tmp_path / "source"
    source_dir.mkdir()
    for name in ("a.jpg", "b.jpg", "c.jpg"):
        PILImage.new("RGB", (40, 40), color="white").save(source_dir / name)

    registry = Registry(
        source_dir=source_dir,
        meta={"provider": {"kind": "claude"}},
        files={
            "a.jpg": FileEntry(status="ok", text="Cher Monsieur, je vous écris depuis Lyon.", model="claude-x"),
            "b.jpg": FileEntry(status="ok", text="Second courrier.", model="claude-x"),
            "c.jpg": FileEntry(status="ok", text="Troisième pièce.", model="claude-x"),
        },
    )
    pieces = build_pieces_from_registry(registry, source_dir)
    piece_a, piece_b, piece_c = pieces
    piece_a.call_number, piece_a.fonds, piece_a.series = "3 U 794/1", "3 U 794", "Correspondance"
    piece_b.call_number, piece_b.fonds, piece_b.series = "3 U 794/2", "3 U 794", "Correspondance"
    piece_c.call_number, piece_c.fonds, piece_c.series = "3 U 794/3", "3 U 794", "Pièces jointes"
    piece_a.confidence_score = 0.95
    piece_b.confidence_score = 0.4
    piece_c.confidence_score = None
    for piece in pieces:
        save_piece(source_dir, piece)
    return source_dir, pieces


def test_pivot_model_documents_field_mapping():
    for field in ("call_number", "date", "sender", "recipient", "transcription"):
        assert "ead" in FIELD_MAPPING[field]
        assert "dublin_core" in FIELD_MAPPING[field]
        assert "mets" in FIELD_MAPPING[field]


def test_field_validation_history_and_mandatory_rejection_note(corpus):
    source_dir, (piece_a, piece_b, _piece_c) = corpus

    for field in ("call_number", "date", "sender", "recipient", "transcription"):
        apply_field_validation(piece_a, field, "validated", changed_by="mb")
    assert global_status(piece_a) == "validated"
    assert len(piece_a.status_history["transcription"]) == 1

    with pytest.raises(ValueError):
        apply_field_validation(piece_b, "transcription", "rejected")
    apply_field_validation(piece_b, "transcription", "rejected", rejection_note="illisible")
    assert piece_b.field_validations["transcription"].rejection_note == "illisible"
    assert global_status(piece_b) == "rejected"


def test_inline_correction_updates_hash_and_flags_stale_translation(corpus):
    source_dir, (piece_a, _piece_b, _piece_c) = corpus
    for field in ("call_number", "date", "sender", "recipient", "transcription"):
        apply_field_validation(piece_a, field, "validated", changed_by="mb")

    provider = _FakeTranslationProvider()
    translate_piece(piece_a, "en", provider, model="fake-model")
    assert piece_a.translations["en"].status == "draft"

    apply_correction(piece_a, "transcription", "Cher Monsieur, texte corrigé après relecture.")
    assert piece_a.translations["en"].status == "stale"


def test_all_twelve_exporters_and_journal_produce_valid_output(tmp_path, corpus):
    _source_dir, pieces = corpus
    piece_a, piece_b, _piece_c = pieces
    for field in ("call_number", "date", "sender", "recipient", "transcription"):
        apply_field_validation(piece_a, field, "validated", changed_by="mb")
    apply_field_validation(piece_b, "call_number", "validated", changed_by="mb")
    translate_piece(piece_a, "en", _FakeTranslationProvider(), model="fake-model")

    out = tmp_path / "out"
    out.mkdir()

    root = ET.parse(export_ead(pieces, out / "ead.xml")).getroot()
    assert len(root.findall(".//c[@level='item']")) == 3

    root = ET.parse(export_dublin_core(pieces, out / "dc.xml")).getroot()
    assert len(list(root)) == 3

    mets_dir = export_mets(pieces, out / "mets")
    assert len(list(mets_dir.glob("*.xml"))) == 3

    csv_path = export_csv_light(pieces, out / "light.csv")
    assert csv_path.exists()

    from regeste.export import export_csv_full

    csv_full_path = export_csv_full(pieces, out / "full.csv")
    assert csv_full_path.exists()

    wb = load_workbook(export_xlsx(pieces, out / "pieces.xlsx"))
    assert wb["Vue d'ensemble"].max_row == 4

    with zipfile.ZipFile(export_zip(pieces, out / "corpus.zip")) as zf:
        assert "metadata.json" in zf.namelist()

    md_path = export_markdown(pieces, out / "corpus.md")
    assert md_path.read_text(encoding="utf-8").count("- **") == 3

    obsidian_dir = export_markdown_obsidian(pieces, out / "obsidian")
    assert len(list(obsidian_dir.glob("*.md"))) == 3

    db_path = export_sqlite(pieces, out / "corpus.db")
    conn = sqlite3.connect(db_path)
    try:
        (count,) = conn.execute("SELECT COUNT(*) FROM pieces").fetchone()
        assert count == 3
    finally:
        conn.close()

    html_path = export_html(pieces, out / "corpus.html")
    assert "<script>" in html_path.read_text(encoding="utf-8")

    pdf_path = export_pdf(pieces, out / "corpus.pdf")
    assert pdf_path.read_bytes().startswith(b"%PDF")

    journal_path = export_review_journal(pieces, out / "journal.xlsx")
    journal_wb = load_workbook(journal_path)
    assert journal_wb["Détail"].max_row == 2 + 3
    summary_rows = {
        journal_wb["Synthèse"].cell(row=r, column=1).value: journal_wb["Synthèse"].cell(row=r, column=2).value
        for r in range(2, journal_wb["Synthèse"].max_row + 1)
    }
    assert summary_rows["3 U 794 - Correspondance"] == 2
    assert summary_rows["3 U 794 - Pièces jointes"] == 1


def test_translation_blocks_and_warns_and_reinjects_glossary_and_entities(corpus):
    _source_dir, (piece_a, piece_b, _piece_c) = corpus

    with pytest.raises(TranslationBlocked):
        translate_piece(piece_b, "en", _FakeTranslationProvider(), model="fake-model")

    for field in ("call_number", "date", "sender", "recipient", "transcription"):
        apply_field_validation(piece_a, field, "validated", changed_by="mb")

    provider = _FakeTranslationProvider()
    translate_piece(piece_a, "en", provider, model="fake-model", glossary={"Lyon": "Lyons"})
    assert "Lyon -> Lyons" in provider.received_prompt
    assert piece_a.transcription in provider.received_prompt


def test_review_queue_sorted_by_confidence_and_bulk_validate(corpus):
    _source_dir, pieces = corpus
    piece_a, piece_b, piece_c = pieces

    ordered = sorted_by_confidence(pieces)
    assert ordered[0] is piece_c  # None confidence reviewed first
    assert ordered[-1] is piece_a  # highest confidence last

    validated = bulk_validate(pieces, threshold=0.5)
    assert validated == [piece_a]
    assert global_status(piece_a) == "validated"
    assert global_status(piece_b) != "validated"  # below threshold, untouched
    assert global_status(piece_c) != "validated"  # unknown confidence, never auto-validated


def test_pivot_persistence_round_trips(corpus):
    source_dir, (piece_a, _piece_b, _piece_c) = corpus
    reloaded = load_piece(source_dir, piece_a.id)
    assert reloaded is not None
    assert reloaded.call_number == piece_a.call_number
    assert reloaded.transcription == piece_a.transcription
