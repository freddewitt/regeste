"""Corpus-level "journal de revue" export — aggregation and fonds/série summary."""

from __future__ import annotations

from openpyxl import load_workbook

from regeste.export import export_review_journal
from regeste.pivot import FieldValidation, Piece


def _corpus():
    validated = Piece(
        id="a.jpg",
        call_number="3 U 794/1",
        fonds="3 U 794",
        series="Correspondance",
        confidence_score=0.95,
        field_validations={
            "call_number": FieldValidation(status="validated", validated_by="mb", validated_at="t1"),
            "date": FieldValidation(status="validated", validated_by="mb", validated_at="t1"),
            "sender": FieldValidation(status="validated", validated_by="mb", validated_at="t1"),
            "recipient": FieldValidation(status="validated", validated_by="mb", validated_at="t1"),
            "transcription": FieldValidation(status="validated", validated_by="mb", validated_at="t1"),
        },
    )
    rejected = Piece(
        id="b.jpg",
        call_number="3 U 794/2",
        fonds="3 U 794",
        series="Correspondance",
        confidence_score=0.3,
        field_validations={
            "transcription": FieldValidation(status="rejected", rejection_note="illisible"),
        },
    )
    draft = Piece(
        id="c.jpg",
        call_number="3 U 794/3",
        fonds="3 U 794",
        series="Pièces jointes",
    )
    return [validated, rejected, draft]


def test_journal_header_has_timestamp(tmp_path):
    path = export_review_journal(_corpus(), tmp_path / "journal.xlsx", generated_at="2026-07-11T00:00:00Z")
    wb = load_workbook(path)
    assert "2026-07-11T00:00:00Z" in wb["Détail"].cell(row=1, column=1).value


def test_journal_detail_has_one_row_per_piece(tmp_path):
    path = export_review_journal(_corpus(), tmp_path / "journal.xlsx")
    wb = load_workbook(path)
    detail = wb["Détail"]
    assert detail.max_row == 2 + 3  # timestamp + header + 3 pieces


def test_journal_summary_aggregates_by_series(tmp_path):
    path = export_review_journal(_corpus(), tmp_path / "journal.xlsx")
    wb = load_workbook(path)
    summary = wb["Synthèse"]
    rows = {
        summary.cell(row=r, column=1).value: summary.cell(row=r, column=2).value
        for r in range(2, summary.max_row + 1)
    }
    assert rows["3 U 794 - Correspondance"] == 2
    assert rows["3 U 794 - Pièces jointes"] == 1


def test_journal_summary_percentages(tmp_path):
    path = export_review_journal(_corpus(), tmp_path / "journal.xlsx")
    wb = load_workbook(path)
    summary = wb["Synthèse"]
    correspondance_row = next(
        r for r in range(2, summary.max_row + 1)
        if summary.cell(row=r, column=1).value == "3 U 794 - Correspondance"
    )
    assert summary.cell(row=correspondance_row, column=3).value == 50.0  # % validé
    assert summary.cell(row=correspondance_row, column=5).value == 50.0  # % rejeté
