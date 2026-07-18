"""Each of the 12 per-piece exporters must produce a valid file on a small test corpus
(>=3 pieces, fonds/série hierarchy, one translated piece, one partially validated piece).
"""

from __future__ import annotations

import csv
import json
import sqlite3
import xml.etree.ElementTree as ET
import zipfile

import pytest
from PIL import Image as PILImage

from regeste.export import (
    export_csv_full,
    export_csv_light,
    export_dublin_core,
    export_ead,
    export_html,
    export_markdown,
    export_markdown_obsidian,
    export_mets,
    export_pdf,
    export_sqlite,
    export_xlsx,
    export_zip,
)
from regeste.pivot import FieldValidation, Piece, Translation


@pytest.fixture
def corpus(tmp_path):
    image_path = tmp_path / "a.jpg"
    PILImage.new("RGB", (40, 40), color="white").save(image_path)

    piece_a = Piece(
        id="a.jpg",
        call_number="3 U 794/1",
        fonds="3 U 794",
        series="Correspondance",
        date="1912",
        sender="X",
        recipient="Y",
        transcription="Cher Monsieur,",
        summary="Lettre anonyme",
        image_path=str(image_path),
        field_validations={
            "call_number": FieldValidation(status="validated"),
            "date": FieldValidation(status="validated"),
            "sender": FieldValidation(status="validated"),
            "recipient": FieldValidation(status="validated"),
            "transcription": FieldValidation(status="validated"),
        },
        translations={"en": Translation(text="Dear Sir,", status="validated", source_hash="h")},
    )
    piece_b = Piece(
        id="b.jpg",
        call_number="3 U 794/2",
        fonds="3 U 794",
        series="Correspondance",
        date="1912",
        sender="X",
        recipient="Z",
        transcription="Second courrier.",
        summary="",
        field_validations={
            "call_number": FieldValidation(status="validated"),
            "date": FieldValidation(status="draft"),
            "sender": FieldValidation(status="draft"),
            "recipient": FieldValidation(status="draft"),
            "transcription": FieldValidation(status="to_review"),
        },
    )
    piece_c = Piece(
        id="c.jpg",
        call_number="3 U 794/3",
        fonds="3 U 794",
        series="Pièces jointes",
        date="1913",
        transcription="Troisième pièce.",
    )
    return [piece_a, piece_b, piece_c]


def test_export_ead_produces_valid_xml(tmp_path, corpus):
    path = export_ead(corpus, tmp_path / "out.xml")
    root = ET.parse(path).getroot()
    assert root.tag == "ead"
    assert len(root.findall(".//c[@level='item']")) == 3


def test_export_dublin_core_produces_valid_xml(tmp_path, corpus):
    path = export_dublin_core(corpus, tmp_path / "out.xml")
    root = ET.parse(path).getroot()
    records = list(root)
    assert len(records) == 3


def test_export_mets_produces_one_valid_document_per_piece(tmp_path, corpus):
    out_dir = export_mets(corpus, tmp_path / "mets")
    files = sorted(out_dir.glob("*.xml"))
    assert len(files) == 3
    for f in files:
        ET.parse(f)  # must not raise


def test_export_csv_light(tmp_path, corpus):
    path = export_csv_light(corpus, tmp_path / "out.csv")
    rows = list(csv.reader(path.open(encoding="utf-8")))
    assert len(rows) == 4  # header + 3


def test_export_csv_full(tmp_path, corpus):
    path = export_csv_full(corpus, tmp_path / "out.csv")
    rows = list(csv.reader(path.open(encoding="utf-8")))
    assert len(rows) == 4
    assert "global_status" in rows[0]


def test_export_xlsx_produces_openable_workbook(tmp_path, corpus):
    from openpyxl import load_workbook

    path = export_xlsx(corpus, tmp_path / "out.xlsx")
    wb = load_workbook(path)
    assert "Vue d'ensemble" in wb.sheetnames
    overview = wb["Vue d'ensemble"]
    assert overview.max_row == 4  # header + 3 pieces


def test_export_zip_contains_expected_layout(tmp_path, corpus):
    path = export_zip(corpus, tmp_path / "out.zip")
    with zipfile.ZipFile(path) as zf:
        names = zf.namelist()
        assert "README.md" in names
        assert "metadata.json" in names
        assert any(n.startswith("transcriptions/") for n in names)
        assert any(n.startswith("images/") for n in names)
        metadata = json.loads(zf.read("metadata.json"))
        assert len(metadata) == 3


def test_export_markdown_simple(tmp_path, corpus):
    path = export_markdown(corpus, tmp_path / "out.md")
    content = path.read_text(encoding="utf-8")
    assert content.count("- **") == 3


def test_export_markdown_obsidian_frontmatter(tmp_path, corpus):
    out_dir = export_markdown_obsidian(corpus, tmp_path / "obsidian")
    files = sorted(out_dir.glob("*.md"))
    assert len(files) == 3
    text = (out_dir / "a.jpg.md").read_text(encoding="utf-8")
    assert text.startswith("---\n")
    assert "statut: validated" in text


def test_export_sqlite_produces_queryable_db(tmp_path, corpus):
    path = export_sqlite(corpus, tmp_path / "out.db")
    conn = sqlite3.connect(path)
    try:
        (count,) = conn.execute("SELECT COUNT(*) FROM pieces").fetchone()
        assert count == 3
        (translation_count,) = conn.execute("SELECT COUNT(*) FROM translations").fetchone()
        assert translation_count == 1
    finally:
        conn.close()


def test_export_html_is_self_contained_and_searchable(tmp_path, corpus):
    path = export_html(corpus, tmp_path / "out.html")
    content = path.read_text(encoding="utf-8")
    assert "<script>" in content
    assert "Cher Monsieur" in content
    assert "http" not in content.split("<script>")[0]  # no external asset in <head>


def test_export_pdf_produces_non_empty_file(tmp_path, corpus):
    path = export_pdf(corpus, tmp_path / "out.pdf")
    assert path.exists()
    assert path.read_bytes().startswith(b"%PDF")


def test_validated_only_filters_pieces(tmp_path, corpus):
    path = export_csv_light(corpus, tmp_path / "out.csv", validated_only=True)
    rows = list(csv.reader(path.open(encoding="utf-8")))
    assert len(rows) == 2  # header + only piece_a (fully validated)


# --- target_language integration (Translation tab, Brique A) ------------------------


def test_csv_light_includes_translation_column_only_when_language_given(tmp_path, corpus):
    without = export_csv_light(corpus, tmp_path / "no_lang.csv")
    rows_without = list(csv.reader(without.open(encoding="utf-8")))
    assert "translation" not in rows_without[0]

    with_lang = export_csv_light(corpus, tmp_path / "en.csv", target_language="en")
    rows = list(csv.reader(with_lang.open(encoding="utf-8")))
    assert rows[0][-1] == "translation"
    assert rows[1][-1] == "Dear Sir,"  # piece_a has an "en" translation
    assert rows[2][-1] == ""  # piece_b has none


def test_csv_full_includes_translation_column(tmp_path, corpus):
    path = export_csv_full(corpus, tmp_path / "out.csv", target_language="en")
    rows = list(csv.reader(path.open(encoding="utf-8")))
    assert rows[0][-1] == "translation"
    assert rows[1][-1] == "Dear Sir,"


def test_xlsx_includes_translation_column(tmp_path, corpus):
    from openpyxl import load_workbook

    path = export_xlsx(corpus, tmp_path / "out.xlsx", target_language="en")
    wb = load_workbook(path)
    overview = wb["Vue d'ensemble"]
    assert overview.cell(row=1, column=10).value == "translation"
    assert overview.cell(row=2, column=10).value == "Dear Sir,"


def test_markdown_simple_includes_translation_when_available(tmp_path, corpus):
    path = export_markdown(corpus, tmp_path / "out.md", target_language="en")
    content = path.read_text(encoding="utf-8")
    assert "Dear Sir," in content


def test_markdown_obsidian_includes_translation_section(tmp_path, corpus):
    out_dir = export_markdown_obsidian(corpus, tmp_path / "obsidian", target_language="en")
    text = (out_dir / "a.jpg.md").read_text(encoding="utf-8")
    assert "## Traduction (en)" in text
    assert "Dear Sir," in text
    other = (out_dir / "b.jpg.md").read_text(encoding="utf-8")
    assert "Traduction" not in other


def test_html_includes_translation_when_available(tmp_path, corpus):
    path = export_html(corpus, tmp_path / "out.html", target_language="en")
    content = path.read_text(encoding="utf-8")
    assert "Dear Sir," in content


def test_pdf_with_target_language_still_produces_valid_file(tmp_path, corpus):
    path = export_pdf(corpus, tmp_path / "out.pdf", target_language="en")
    assert path.read_bytes().startswith(b"%PDF")


def test_sqlite_pieces_table_gets_translation_column(tmp_path, corpus):
    path = export_sqlite(corpus, tmp_path / "out.db", target_language="en")
    conn = sqlite3.connect(path)
    try:
        (translation,) = conn.execute(
            "SELECT translation FROM pieces WHERE id = 'a.jpg'"
        ).fetchone()
        assert translation == "Dear Sir,"
    finally:
        conn.close()


def test_dc_ead_mets_list_available_languages(tmp_path, corpus):
    # piece_a: language_detected="" (not set in fixture) + translation "en" -> just "en"
    dc_path = export_dublin_core(corpus, tmp_path / "dc.xml")
    root = ET.parse(dc_path).getroot()
    dc_ns = "{http://purl.org/dc/elements/1.1/}"
    record_a = root[0]
    languages = [el.text for el in record_a.findall(f"{dc_ns}language")]
    assert languages == ["en"]

    ead_path = export_ead(corpus, tmp_path / "ead.xml")
    ead_root = ET.parse(ead_path).getroot()
    lang_elements = ead_root.findall(".//language")
    assert any(el.get("langcode") == "en" for el in lang_elements)

    mets_dir = export_mets(corpus, tmp_path / "mets")
    mets_root = ET.parse(mets_dir / "a.jpg.xml").getroot()
    mets_ns = "{http://www.loc.gov/METS/}"
    languages_mets = [el.text for el in mets_root.findall(f".//{mets_ns}language")]
    assert languages_mets == ["en"]
